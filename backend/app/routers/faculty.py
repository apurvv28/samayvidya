"""Faculty management routes."""
import csv
import io
import re
from typing import Any

from fastapi import APIRouter, HTTPException, status, Depends, UploadFile, File
from openpyxl import load_workbook
from pydantic import BaseModel
from app.dependencies.auth import get_current_user, CurrentUser
from app.supabase_client import get_user_supabase, get_service_supabase
from app.schemas.common import SuccessResponse, FacultyRoleEnum, SubjectTypeEnum
from app.services.email_service import send_faculty_credentials

router = APIRouter(prefix="/faculty", tags=["faculty"])


class FacultyCreate(BaseModel):
    """Create faculty request."""

    faculty_code: str
    faculty_name: str
    role: FacultyRoleEnum
    priority_level: int
    preferred_start_time: str
    preferred_end_time: str
    min_working_days: int
    max_working_days: int
    max_load_per_week: int
    department_id: str
    is_active: bool = True
    designation: str | None = None
    email: str | None = None
    phone: str | None = None
    target_theory_load: int = 0
    target_lab_load: int = 0
    target_tutorial_load: int = 0
    target_other_load: int = 0


class FacultyUpdate(BaseModel):
    """Update faculty request."""

    faculty_code: str | None = None
    faculty_name: str | None = None
    role: FacultyRoleEnum | None = None
    priority_level: int | None = None
    preferred_start_time: str | None = None
    preferred_end_time: str | None = None
    min_working_days: int | None = None
    max_working_days: int | None = None
    max_load_per_week: int | None = None
    is_active: bool | None = None
    designation: str | None = None
    target_theory_load: int | None = None
    target_lab_load: int | None = None
    target_tutorial_load: int | None = None
    target_other_load: int | None = None


CSV_HEADER_ALIASES: dict[str, list[str]] = {
    "faculty_code": ["faculty_code", "faculty code", "faculty id", "code"],
    "faculty_name": ["faculty_name", "faculty name", "name", "full name"],
    "email": ["email", "email id", "mail"],
    "phone": ["phone", "phone number", "mobile", "contact"],
    "designation": ["designation", "title"],
    "department_id": ["department_id", "department id", "dept_id"],
    "department_name": ["department_name", "department name", "department", "dept"],
    "role": ["role"],
    "priority_level": ["priority_level", "priority level", "priority"],
    "max_load_per_week": ["max_load_per_week", "max load per week", "max load"],
    "preferred_start_time": ["preferred_start_time", "preferred start time", "start time"],
    "preferred_end_time": ["preferred_end_time", "preferred end time", "end time"],
    "min_working_days": ["min_working_days", "min working days"],
    "max_working_days": ["max_working_days", "max working days"],
    "is_active": ["is_active", "active"],
    "target_theory_load": ["target_theory_load", "target theory load", "theory load"],
    "target_lab_load": ["target_lab_load", "target lab load", "lab load"],
    "target_tutorial_load": ["target_tutorial_load", "target tutorial load", "tutorial load"],
    "target_other_load": ["target_other_load", "target other load", "other load"],
}


def _normalize_header(value: str) -> str:
    return " ".join(value.strip().lower().replace("_", " ").split())


def _find_header_key(normalized_headers: dict[str, str], aliases: list[str]) -> str | None:
    for alias in aliases:
        key = normalized_headers.get(_normalize_header(alias))
        if key:
            return key
    return None


def _clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _extract_department_hint(text: str | None) -> str | None:
    if not text:
        return None
    match = re.search(r"department\s*[-:]\s*([a-z0-9&/ -]+)", text, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1).strip()


def _extract_sheet_candidate(worksheet: Any) -> tuple[list[dict[str, str]], list[str], str | None] | None:
    max_columns = worksheet.max_column
    max_scan_rows = min(worksheet.max_row, 40)
    header_row_index: int | None = None

    for row_idx in range(1, max_scan_rows + 1):
        row_values = [
            _clean_cell_value(worksheet.cell(row_idx, col_idx).value)
            for col_idx in range(1, max_columns + 1)
        ]
        normalized_row = {_normalize_header(value) for value in row_values if value}
        if "faculty name" in normalized_row and (
            "designation" in normalized_row
            or "sr.no." in normalized_row
            or "sr no." in normalized_row
            or "faculty load (hours)" in normalized_row
        ):
            header_row_index = row_idx
            break

    if header_row_index is None:
        return None

    department_hint = None
    for row_idx in range(1, header_row_index):
        cell_text = _clean_cell_value(worksheet.cell(row_idx, 1).value)
        candidate = _extract_department_hint(cell_text)
        if candidate:
            department_hint = candidate
            break

    headers: list[str] = []
    for col_idx in range(1, max_columns + 1):
        header = _clean_cell_value(worksheet.cell(header_row_index, col_idx).value) or f"column_{col_idx}"
        unique_header = header
        suffix = 2
        while unique_header in headers:
            unique_header = f"{header}_{suffix}"
            suffix += 1
        headers.append(unique_header)

    rows: list[dict[str, str]] = []
    for row_idx in range(header_row_index + 1, worksheet.max_row + 1):
        row_values = [
            _clean_cell_value(worksheet.cell(row_idx, col_idx).value)
            for col_idx in range(1, max_columns + 1)
        ]
        if not any(row_values):
            continue
        rows.append({headers[col_idx - 1]: row_values[col_idx - 1] for col_idx in range(1, max_columns + 1)})

    if not rows:
        return None

    return rows, headers, department_hint


def _read_xlsx_rows(content: bytes) -> tuple[list[dict[str, str]], list[str], str | None]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    best_candidate: tuple[list[dict[str, str]], list[str], str | None] | None = None
    best_score = (-1, -1)

    for worksheet in workbook.worksheets:
        candidate = _extract_sheet_candidate(worksheet)
        if not candidate:
            continue

        rows, headers, department_hint = candidate
        normalized_headers = {_normalize_header(header): header for header in headers}
        faculty_name_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["faculty_name"])
        non_empty_names = 0
        if faculty_name_key:
            non_empty_names = sum(
                1
                for row in rows
                if _clean_cell_value(row.get(faculty_name_key))
            )

        score = (non_empty_names, len(rows))
        if score > best_score:
            best_score = score
            best_candidate = candidate

    if not best_candidate:
        raise ValueError("No valid faculty sheet found in XLSX file.")

    return best_candidate


def _generate_faculty_code(name: str, row_number: int) -> str:
    cleaned = "".join(char for char in name.upper() if char.isalnum())
    prefix = cleaned[:6] if cleaned else "FAC"
    return f"{prefix}{row_number:03d}"


def _generate_faculty_email(name: str, row_number: int) -> str:
    slug = re.sub(r"[^a-z0-9]+", ".", name.lower()).strip(".")
    if not slug:
        slug = f"faculty{row_number}"
    return f"{slug}.{row_number}@noemail.local"


def _autofill_excel_required_fields(
    rows: list[dict[str, str]],
    fieldnames: list[str],
    department_hint: str | None,
) -> tuple[list[dict[str, str]], list[str]]:
    normalized_headers = {_normalize_header(header): header for header in fieldnames}
    name_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["faculty_name"])
    code_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["faculty_code"])
    email_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["email"])
    department_id_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["department_id"])
    department_name_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["department_name"])
    role_key = _find_header_key(normalized_headers, CSV_HEADER_ALIASES["role"])

    if not name_key:
        return rows, fieldnames

    generated_code_key = code_key or "Faculty Code"
    generated_email_key = email_key or "Email"
    generated_department_key = department_name_key or "Department Name"
    generated_role_key = role_key or "Role"

    if not code_key:
        fieldnames.append(generated_code_key)
    if not email_key:
        fieldnames.append(generated_email_key)
    if not role_key:
        fieldnames.append(generated_role_key)
    if not department_id_key and not department_name_key and department_hint:
        fieldnames.append(generated_department_key)

    for row_index, row in enumerate(rows, start=1):
        faculty_name = _clean_cell_value(row.get(name_key))
        if not faculty_name:
            continue

        if not code_key:
            row[generated_code_key] = _generate_faculty_code(faculty_name, row_index)
        if not email_key:
            row[generated_email_key] = _generate_faculty_email(faculty_name, row_index)
        if not role_key:
            row[generated_role_key] = "FACULTY"
        if not department_id_key and not department_name_key and department_hint:
            row[generated_department_key] = department_hint

    return rows, fieldnames


def _get_csv_value(row: dict[str, Any], normalized_headers: dict[str, str], aliases: list[str]) -> str | None:
    for alias in aliases:
        key = normalized_headers.get(_normalize_header(alias))
        if not key:
            continue
        value = row.get(key)
        if value is None:
            continue
        cleaned = str(value).strip()
        if cleaned:
            return cleaned
    return None


def _parse_int(value: str | None, default: int) -> int:
    if value is None:
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid integer value '{value}'") from exc


def _parse_bool(value: str | None, default: bool = True) -> bool:
    if value is None:
        return default
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y"}:
        return True
    if normalized in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Invalid boolean value '{value}'")


def _derive_priority_and_max_load(designation: str | None) -> tuple[int, int]:
    normalized = (designation or "").strip().lower()
    if normalized == "professor":
        return 1, 14
    if normalized == "associate professor":
        return 2, 18
    if normalized == "assistant professor":
        return 3, 20
    return 4, 22


def _create_faculty_with_auth(supabase: Any, faculty: FacultyCreate) -> tuple[Any, str]:
    existing_faculty = (
        supabase.table("faculty")
        .select("faculty_id")
        .eq("faculty_code", faculty.faculty_code)
        .execute()
    )
    if existing_faculty.data:
        raise HTTPException(
            status_code=400,
            detail=f"Faculty code '{faculty.faculty_code}' already exists.",
        )

    if not faculty.email:
        raise HTTPException(status_code=400, detail="Email is required for new faculty")

    first_name = faculty.faculty_name.split()[0]
    phone_digits = "".join(filter(str.isdigit, faculty.phone or ""))
    phone_suffix = phone_digits[-4:] if len(phone_digits) >= 4 else "1234"
    password = f"{first_name}{phone_suffix}"

    try:
        user_attributes = {
            "email": faculty.email,
            "password": password,
            "email_confirm": True,
            "user_metadata": {
                "full_name": faculty.faculty_name,
                "role": "FACULTY",
                "designation": faculty.designation or "",
            },
        }
        auth_response = supabase.auth.admin.create_user(user_attributes)
        user_id = auth_response.user.id
    except Exception as auth_error:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to create user account: {str(auth_error)}",
        ) from auth_error

    try:
        profile_data = {
            "id": user_id,
            "email": faculty.email,
            "full_name": faculty.faculty_name,
            "role": "FACULTY",
            "is_active": True,
        }
        supabase.table("user_profiles").upsert(profile_data).execute()
    except Exception as profile_error:
        print(f"Profile creation warning: {profile_error}")

    faculty_data = faculty.model_dump()
    response = supabase.table("faculty").insert(faculty_data).execute()

    message = "Faculty created successfully"
    if faculty.email and faculty.email.endswith("@noemail.local"):
        email_sent = False
        message += " (auto-generated email; credential email skipped)."
    else:
        email_sent = send_faculty_credentials(
            to_email=faculty.email,
            name=faculty.faculty_name,
            password=password,
            faculty_id=faculty.faculty_code,
        )
        if not email_sent:
            message += ", but failed to send email."
        else:
            message += " and email sent."

    return response.data, message


def _build_faculty_from_csv_row(
    row: dict[str, Any],
    normalized_headers: dict[str, str],
    department_name_to_id: dict[str, str],
    valid_department_ids: set[str],
) -> FacultyCreate:
    faculty_code = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["faculty_code"])
    faculty_name = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["faculty_name"])
    email = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["email"])
    phone = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["phone"])
    designation = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["designation"]) or "Assistant Professor"
    role_raw = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["role"]) or "FACULTY"

    department_name = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["department_name"])
    department_id = _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["department_id"])
    if department_name:
        normalized_department_name = _normalize_header(department_name)
        fallback_department_id = department_name_to_id.get(normalized_department_name)
        if not fallback_department_id and "-" in normalized_department_name:
            for token in [part.strip() for part in normalized_department_name.split("-") if part.strip()]:
                fallback_department_id = department_name_to_id.get(token)
                if fallback_department_id:
                    break
        if (
            not fallback_department_id
            and normalized_department_name
        ):
            for known_department_name, known_department_id in department_name_to_id.items():
                if known_department_name in normalized_department_name or normalized_department_name in known_department_name:
                    fallback_department_id = known_department_id
                    break
    else:
        fallback_department_id = None

    if department_id and department_id not in valid_department_ids and fallback_department_id:
        department_id = fallback_department_id
    elif not department_id and fallback_department_id:
        department_id = fallback_department_id

    if not faculty_code:
        raise ValueError("Missing faculty code")
    if not faculty_name:
        raise ValueError("Missing faculty name")
    if not email:
        raise ValueError("Missing email")
    if not department_id:
        raise ValueError("Missing department (use Department ID or Department Name)")
    if department_id not in valid_department_ids:
        raise ValueError(f"Invalid department id '{department_id}'")

    try:
        role = FacultyRoleEnum(role_raw.strip().upper().replace(" ", "_"))
    except ValueError as exc:
        allowed_roles = ", ".join(member.value for member in FacultyRoleEnum)
        raise ValueError(f"Invalid role '{role_raw}'. Allowed: {allowed_roles}") from exc

    derived_priority, derived_max_load = _derive_priority_and_max_load(designation)
    priority_level = _parse_int(
        _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["priority_level"]),
        derived_priority,
    )
    max_load_per_week = _parse_int(
        _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["max_load_per_week"]),
        derived_max_load,
    )

    return FacultyCreate(
        faculty_code=faculty_code,
        faculty_name=faculty_name,
        role=role,
        priority_level=priority_level,
        preferred_start_time=_get_csv_value(
            row,
            normalized_headers,
            CSV_HEADER_ALIASES["preferred_start_time"],
        )
        or "09:00",
        preferred_end_time=_get_csv_value(
            row,
            normalized_headers,
            CSV_HEADER_ALIASES["preferred_end_time"],
        )
        or "17:00",
        min_working_days=_parse_int(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["min_working_days"]),
            5,
        ),
        max_working_days=_parse_int(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["max_working_days"]),
            6,
        ),
        max_load_per_week=max_load_per_week,
        department_id=department_id,
        is_active=_parse_bool(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["is_active"]),
            True,
        ),
        designation=designation,
        email=email,
        phone=phone,
        target_theory_load=_parse_int(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["target_theory_load"]),
            0,
        ),
        target_lab_load=_parse_int(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["target_lab_load"]),
            0,
        ),
        target_tutorial_load=_parse_int(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["target_tutorial_load"]),
            0,
        ),
        target_other_load=_parse_int(
            _get_csv_value(row, normalized_headers, CSV_HEADER_ALIASES["target_other_load"]),
            0,
        ),
    )


@router.get("", response_model=SuccessResponse)
async def list_faculty(
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """List all faculty members (Service Role - Bypasses RLS)."""
    try:
        supabase = get_service_supabase()
        response = supabase.table("faculty").select("*").execute()
        return {"data": response.data, "message": "Faculty retrieved successfully"}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch faculty: {str(e)}",
        )


@router.post("/upload", response_model=SuccessResponse)
async def upload_faculty_csv(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Upload and process faculty CSV in bulk."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required.")

    file_name = file.filename.lower()
    if not (file_name.endswith(".csv") or file_name.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload a CSV or XLSX file.")

    try:
        content = await file.read()
        department_hint = None
        if file_name.endswith(".csv"):
            decoded_content = content.decode("utf-8-sig")
            csv_reader = csv.DictReader(io.StringIO(decoded_content))
            if not csv_reader.fieldnames:
                raise HTTPException(status_code=400, detail="CSV file has no headers.")
            rows = list(csv_reader)
            fieldnames = [header for header in csv_reader.fieldnames if header and header.strip()]
        else:
            rows, fieldnames, department_hint = _read_xlsx_rows(content)
            rows, fieldnames = _autofill_excel_required_fields(rows, fieldnames, department_hint)

        normalized_headers = {
            _normalize_header(header): header
            for header in fieldnames
            if header and header.strip()
        }

        def has_alias(column_key: str) -> bool:
            for alias in CSV_HEADER_ALIASES[column_key]:
                if _normalize_header(alias) in normalized_headers:
                    return True
            return False

        required_columns = ["faculty_code", "faculty_name", "email"]
        missing_columns = [column for column in required_columns if not has_alias(column)]
        has_department = has_alias("department_id") or has_alias("department_name")
        if not has_department:
            missing_columns.append("department_id/department_name")
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required CSV columns: {', '.join(missing_columns)}",
            )

        if not rows:
            raise HTTPException(status_code=400, detail="Uploaded file has no data rows.")

        supabase = get_service_supabase()
        department_response = supabase.table("departments").select("department_id, department_name").execute()
        department_name_to_id = {
            _normalize_header(department["department_name"]): department["department_id"]
            for department in (department_response.data or [])
            if department.get("department_name") and department.get("department_id")
        }
        valid_department_ids = {
            department["department_id"]
            for department in (department_response.data or [])
            if department.get("department_id")
        }

        results = {
            "total_rows": len(rows),
            "created": 0,
            "failed": 0,
            "errors": [],
            "created_faculty": [],
        }

        for row_number, row in enumerate(rows, start=2):
            try:
                faculty = _build_faculty_from_csv_row(
                    row,
                    normalized_headers,
                    department_name_to_id,
                    valid_department_ids,
                )
                created_data, _ = _create_faculty_with_auth(supabase, faculty)
                results["created"] += 1

                if isinstance(created_data, list) and created_data:
                    faculty_row = created_data[0]
                else:
                    faculty_row = created_data or {}

                results["created_faculty"].append(
                    {
                        "faculty_id": faculty_row.get("faculty_id"),
                        "faculty_code": faculty_row.get("faculty_code", faculty.faculty_code),
                        "faculty_name": faculty_row.get("faculty_name", faculty.faculty_name),
                    }
                )
            except (HTTPException, ValueError) as error:
                results["failed"] += 1
                if isinstance(error, HTTPException):
                    detail = error.detail
                else:
                    detail = str(error)
                results["errors"].append(f"Row {row_number}: {detail}")
            except Exception as error:
                results["failed"] += 1
                results["errors"].append(f"Row {row_number}: {str(error)}")

        return {
            "data": results,
            "message": (
                f"Processed {results['total_rows']} rows. "
                f"Created: {results['created']}, Failed: {results['failed']}."
            ),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"CSV processing failed: {str(e)}",
        )


@router.get("/{faculty_id}", response_model=SuccessResponse)
async def get_faculty(
    faculty_id: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Get a specific faculty member by ID."""
    try:
        supabase = get_user_supabase()
        response = (
            supabase.table("faculty")
            .select("*")
            .eq("faculty_id", faculty_id)
            .single()
            .execute()
        )
        return {
            "data": response.data,
            "message": "Faculty retrieved successfully",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Faculty not found: {str(e)}",
        )


@router.post("", response_model=SuccessResponse)
async def create_faculty(
    faculty: FacultyCreate,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Create a new faculty member with Auth user and Email notification."""
    try:
        supabase = get_service_supabase()
        response_data, message = _create_faculty_with_auth(supabase, faculty)
        return {
            "data": response_data,
            "message": message,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to create faculty: {str(e)}",
        )


@router.put("/{faculty_id}", response_model=SuccessResponse)
async def update_faculty(
    faculty_id: str,
    faculty: FacultyUpdate,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Update a faculty member."""
    try:
        supabase = get_service_supabase()
        update_data = faculty.model_dump(exclude_unset=True)
        response = (
            supabase.table("faculty")
            .update(update_data)
            .eq("faculty_id", faculty_id)
            .execute()
        )
        return {
            "data": response.data,
            "message": "Faculty updated successfully",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to update faculty: {str(e)}",
        )


@router.delete("/{faculty_id}", response_model=SuccessResponse)
async def delete_faculty(
    faculty_id: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Delete a faculty member."""
    try:
        supabase = get_user_supabase()
        response = (
            supabase.table("faculty")
            .delete()
            .eq("faculty_id", faculty_id)
            .execute()
        )
        return {
            "data": response.data,
            "message": "Faculty deleted successfully",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to delete faculty: {str(e)}",
        )



class FacultySubjectAssign(BaseModel):
    subject_id: str
    division_id: str | None = None
    batch_id: str | None = None
    session_type: SubjectTypeEnum


@router.get("/{faculty_id}/subjects", response_model=SuccessResponse)
async def get_faculty_subjects(
    faculty_id: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Get subjects assigned to a faculty member."""
    try:
        supabase = get_user_supabase()
        # Join with subjects table and other related tables
        # Adjusting query to fetch related data from the new mapping table
        # Note: We need to ensure foreign keys are correctly set up in Supabase for this join to work seamlessly with `select` syntax.
        # Assuming `subjects`, `divisions`, `batches` are related.
        response = (
            supabase.table("faculty_subject_mapping")
            .select("*, subjects(*), divisions(*), batches(*)")
            .eq("faculty_id", faculty_id)
            .execute()
        )
        return {
            "data": response.data,
            "message": "Faculty subjects retrieved successfully",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch faculty subjects: {str(e)}",
        )


@router.post("/{faculty_id}/subjects", response_model=SuccessResponse)
async def assign_subject_to_faculty(
    faculty_id: str,
    assignment: FacultySubjectAssign,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Assign a subject to a faculty member."""
    try:
        supabase = get_service_supabase()
        
        # Verify uniqueness or other business logic if needed (e.g. check overlaps)
        # For now, allowing multiple assignments as per schema unless unique constraint violations occur.

        data = {
            "faculty_id": faculty_id,
            "subject_id": assignment.subject_id,
            "division_id": assignment.division_id,
            "batch_id": assignment.batch_id,
            "session_type": assignment.session_type
        }
        
        response = (
            supabase.table("faculty_subject_mapping")
            .insert(data)
            .execute()
        )
        return {
            "data": response.data,
            "message": "Subject assigned to faculty successfully",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to assign subject: {str(e)}",
        )

@router.delete("/{faculty_id}/subjects/{mapping_id}", response_model=SuccessResponse)
async def unassign_subject_from_faculty(
    faculty_id: str,
    mapping_id: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Unassign a subject from a faculty member (Delete Mapping)."""
    try:
        supabase = get_service_supabase()
        response = (
            supabase.table("faculty_subject_mapping")
            .delete()
            .eq("mapping_id", mapping_id)
            .eq("faculty_id", faculty_id) # Extra safety check
            .execute()
        )
        return {
            "data": response.data,
            "message": "Subject unassigned from faculty successfully",
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to unassign subject: {str(e)}",
        )
